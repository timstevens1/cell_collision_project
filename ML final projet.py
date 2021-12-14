import random

import numpy as np
import pandas as pd
import sklearn.metrics
import tensorflow as tf
import openpyxl as op

from tensorflow.keras.models import Sequential
from tensorflow.keras.layers import Dense, Flatten
from tensorflow.keras import optimizers
import plotly.express as px

import matplotlib.pyplot as plt
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
from tensorflow.keras.utils import to_categorical

"""
wb = op.load_workbook(data)
sheet = wb.active

for row in range(2, sheet.max_row + 1):
    name = sheet.cell(row, 3).value
    if 'Mono' in name:
        sheet.cell(row, 3).value = 1
    elif 'Dead' in name:
        sheet.cell(row, 3).value = 0
    elif 'Uninf' in name:
        sheet.cell(row, 3).value = 2
    elif 'Syn' in name:
        sheet.cell(row, 3).value = 3
    else:
        sheet.cell(row, 3).value = 4


wb.save('Filtered tracks.xlsx')
mono = 1
dead = 0
syn = 3
unif = 2
else = 4
"""

# import the data
tracks = "Filtered tracks.xlsx"
results = 'Huge_V6_results.xlsx'
results = pd.read_excel(results, sheet_name='Summary')
tracks = pd.read_excel(tracks)
df_tracks = pd.DataFrame(tracks)
df = pd.DataFrame(results)
df['Type'] = df_tracks['Tags']
print(df.shape)

df = df.drop(['Velocity Mean (Filtered)', 'Velocity Median (Filtered)', 'Velocity SD (Filtered)',
              'Acceleration Mean (Filtered)', 'Acceleration Median (Filtered)', 'Acceleration SD (Filtered)',
              'Angle LOBF Slope', 'Angle LOBF R2', 'Duration'], axis=1)

df_without_dead = df[df['Arrest Coefficient'] <= 0.5]
df_uninf = df_without_dead[df_without_dead['Type'] == 2]
df_half = df_uninf.iloc[:317, :]
frames = [df_half, df_without_dead]
df = pd.concat(frames, copy=False)
df = df[df['Arrest Coefficient'] <= 0.5]
print(df['Type'])
print(df_uninf.shape, df.shape, df_half.shape, df_without_dead.shape)

# split into x an y
# without_dead = df.loc[df['Type of Cell'] == 1 | 3 | 4]

df = df.fillna(180)
df.head()
x = df.iloc[:, :19]
y = df.iloc[:, 19:]
x = StandardScaler().fit_transform(x)
y = StandardScaler().fit_transform(y)

# need to think how to regularize and how to also account for missing values for angles

x_train, x_test, y_train, y_test = train_test_split(x, y, train_size=0.90, random_state=0)

print('x_train ', x_train.shape, '\n', 'y_train ', y_train.shape, '\n', 'x_test ', x_test.shape, '\n', 'y_test',
      y_test.shape)
y_test = to_categorical(y_test, num_classes=4)
y_train = to_categorical(y_train, num_classes=4)

df_final = pd.DataFrame(columns=['Batch Size', 'Learning Rate', 'Accuracy'])


def nn(y_test):
    layers = [
        Flatten(),
        Dense(100, activation='relu'),  # adds layer
        Dense(100, activation='relu'),
        Dense(100, activation='relu'),
        Dense(100, activation='relu'),
        Dense(4, activation='softmax'),  # 4 since we have 4 classes in the dataset.
    ]
    model = Sequential(layers)

    model.compile(optimizer=tf.keras.optimizers.Adam(learning_rate=0.0002),
                  loss='categorical_crossentropy',
                  metrics=['accuracy'])  # look into confusion matrix

    history_of_performance = model.fit(x_train, y_train, batch_size=15, epochs=30, validation_split=0.20)
    y_pred = model.predict(x_test, batch_size=15)
    y_pred = np.argmax(y_pred, axis=1)
    y_test = np.argmax(y_test, axis=1)
    print(classification_report(y_true=y_test, y_pred=y_pred, output_dict=True))
    print(confusion_matrix(y_true=y_test, y_pred=y_pred))


nn(y_test)

